import openpyxl



def RowCount(file, sheet_name):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    return sheet.max_row

def ColumnCount(file, sheet_name):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    return sheet.max_column

def ReadData(file, sheet_name, row_num, col_num):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    return sheet.cell(row=row_num, column=col_num).value

def WriteData(file, sheet_name, row_num, col_num, data):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    sheet.cell(row=row_num, column=col_num).value = data
    book.save(file)